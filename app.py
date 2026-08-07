from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.utils import get_column_letter

# --- داخل جزء تحميل التقرير (بدل الكود القديم للتحميل) ---
output = io.BytesIO()
with pd.ExcelWriter(output, engine='openpyxl') as writer:
    df_export.to_excel(writer, index=False, sheet_name='Report')
    
    # الوصول إلى ورقة العمل لتعديل التنسيق
    workbook = writer.book
    worksheet = writer.sheets['Report']
    
    # جعل الورقة من اليمين إلى اليسار
    worksheet.sheet_view.rightToLeft = True
    
    # تعريف الألوان (مثل تصميم التطبيق)
    header_fill = PatternFill(start_color="004D40", end_color="004D40", fill_type="solid")
    header_font = Font(color="D4AF37", bold=True)
    
    # تنسيق رأس الجدول
    for col_num, value in enumerate(df_export.columns.values, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        # توسيع الأعمدة تلقائياً
        worksheet.column_dimensions[get_column_letter(col_num)].width = 20

    # تلوين الصفوف بالتناوب (أخضر فاتح وأصفر فاتح)
    row_fill_odd = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    row_fill_even = PatternFill(start_color="FEFDE8", end_color="FEFDE8", fill_type="solid")
    
    for row_num in range(2, worksheet.max_row + 1):
        fill = row_fill_odd if row_num % 2 == 0 else row_fill_even
        for col_num in range(1, worksheet.max_column + 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.fill = fill
            cell.alignment = Alignment(horizontal="right")

excel_data = output.getvalue()

st.download_button(
    label="📥 تحميل التقرير المنسق (xlsx)",
    data=excel_data,
    file_name=f"mosque_report_{datetime.now().strftime('%Y%m%d')}.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)
